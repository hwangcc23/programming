#!/usr/bin/python
# This Python file uses the following encoding: utf-8

"""
gen_CR_review_tbl helps me automate a fucking boring stuff: parse a given excel file from CQ,
and translate it into a specific format for my hosting issue review meeting.

Copyright (C) 2017 Chih-Chyuan Hwang (hwangcc@csie.nctu.edu.tw)

This program is free software; you can redistribute it and/or modify
it under the terms of the GNU General Public License version 2 as
published by the Free Software Foundation.
"""

import sys
import getopt
import openpyxl
import copy
from openpyxl.styles import NamedStyle, Color, colors, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

CON_FIND_ASSIGNEE = "find_assignee"
CON_BYPASS_MODEM = "bypass_modem"
CON_BYPASS_CONN = "bypass_conn"
CON_BYPASS_MM = "bypass_mm"
CON_BYPASS_QA = "bypass_qa"
CON_SORTING = "sorting"
CON_REMOVE_APPIOT_CR = "remove_appiot_cr"

sorting_titles = ["id", "Severity", "Assignee.groups.name", "Assignee_Name"]

def usage():
    print("gen_CR_review_tbl: Gnerate the CR review table from the raw CQ excel file")
    print("Usage: gen_CR_review_tbl [options]")
    print("       options and arguments:")
    print("       -h: show help")
    print("       -i|--input FILENAME: give the raw CQ excel file")
    print("       -o|--output FILENAME: specify the output file name")
    print("       -m|--mapping FILENAME: give the team/window mapping file")
    print("       -c|--condition CONDITIONS: give the extra conditions(supported conditions: find_assignee,sorting,bypass_modem,bypass_mm,bypass_conn,bypass_qa,remove_appiot_cr)")

def team_window_mapping(mapping_file):
    wb = openpyxl.load_workbook(mapping_file)
    sheet = wb.active

    col_team = 0
    col_window = 0
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == "SW Contact Window":
            col_window = col
        elif sheet.cell(row=1, column=col).value == "SW Team":
            col_team = col
    if col_team == 0 or col_window == 0:
        print("The mapping file " + mapping_file + " is invalid")
        if col_window == 0:
            print("Columns of [SW Contact Winsow] cannot be found")
        if col_team == 0:
            print("Columns of [SW Team] cannot be found")
        sys.exit(0)

    team_windows = []
    for row in range(2, sheet.max_row + 1):
        team_window = {}
        team_window["team"] = sheet.cell(row=row, column=col_team).value
        team_window["window"] = sheet.cell(row=row, column=col_window).value
        team_windows.append(copy.copy(team_window))

    return team_windows

def team_category(team):
    if team.upper().find("QA") != -1 or team.upper().find("VEND_") != -1 \
    or team.upper().find("VENDOR_") != -1 or team == "":
        return "QA"
    elif team.upper().find("WSP") != -1 or team.upper().find("WCS") != -1 \
    or team.upper().find("WCT") != -1 or team.upper().find("CSD") != -1:
        return "Modem"
    elif team.upper().find("CTD") != -1 or team.upper().find("WSD_SE") != -1:
        return "Conn"
    elif team.upper().find("MM") != -1:
        return "MM"
    else:
        return "AP"

def gen_CR_review_tbl(input_file, output_file, mapping_file, condition):
    print("input_file = " + input_file, ", output_file = " + output_file)
    print("Generate CR review table...")
    print("")

    CRs = []
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    nr_appiot_cr = 0
    for row in range(2, sheet.max_row + 1):
        CR = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            CR[sheet.cell(row=1, column=col).value] = cell.value
        if condition.find(CON_REMOVE_APPIOT_CR) != -1:
            if 'Sqa_Feature_Group' in CR and CR["Sqa_Feature_Group"].find("APPIOT") != -1:
                nr_appiot_cr += 1
                continue
            elif 'Feature_Name' in CR and CR["Feature_Name"].find("APP IOT") != -1:
                nr_appiot_cr += 1
                continue
            elif 'Title' in CR and CR["Title"].find("APPIOT") != -1:
                nr_appiot_cr += 1
                continue
        CRs.append(copy.copy(CR))
    print("Total number of counting CRs: %d" % (len(CRs) + nr_appiot_cr))
    if condition.find(CON_REMOVE_APPIOT_CR) != -1:
        print("Remove %d APPIOT CRs" % (nr_appiot_cr))
    #print(CRs)
    #for i in range(0, len(CRs)):
    #    print("id =", CRs[i]["id"], "Severity =", CRs[i]["Severity"], "Assignee.groups =", CRs[i]["Assignee.groups"])

    team_windows = team_window_mapping(mapping_file)
    #print(team_windows)

    review_tbl = []
    for i in range(0, len(CRs)):
        existing = 0
        for j in range(0, len(review_tbl)):
            if ("Assignee.groups.name" in CRs[i] and CRs[i]["Assignee.groups.name"] == review_tbl[j]["team"]) \
            or ("Assignee.groups" in CRs[i] and CRs[i]["Assignee.groups"] == review_tbl[j]["team"]):
                review_tbl[j]["count"] += 1
                if condition.find(CON_FIND_ASSIGNEE) != -1 or review_tbl[j]["have_window"] == 0:
                    if review_tbl[j]["window"].find(CRs[i]["Assignee_Name"]) == -1:
                        review_tbl[j]["window"] += ";" + CRs[i]["Assignee_Name"]
                existing = 1
        if existing == 0:
            review_rec = {}
            if "Assignee.groups.name" in CRs[i]:
                review_rec["team"] = CRs[i]["Assignee.groups.name"]
            elif "Assignee.groups" in CRs[i]:
                review_rec["team"] = CRs[i]["Assignee.groups"]
            review_rec["category"] = team_category(review_rec["team"])
            review_rec["count"] = 1
            review_rec["have_window"] = 0
            for k in range(0, len(team_windows)):
                if review_rec["team"].upper() == team_windows[k]["team"].upper():
                    review_rec["window"] = team_windows[k]["window"]
                    review_rec["have_window"] = 1
                    break
            if condition.find(CON_FIND_ASSIGNEE) != -1:
                review_rec["window"] = CRs[i]["Assignee_Name"]
            # NoteXXX: If the team window cannot be found, assign the assignee
            elif review_rec["have_window"] == 0:
                review_rec["window"] = CRs[i]["Assignee_Name"]
            review_tbl.append(copy.copy(review_rec))
    review_tbl = sorted(review_tbl, key=lambda x: x["category"])
    #for i in range(0, len(review_tbl)):
    #    print(review_tbl[i]["team"], review_tbl[i]["count"], review_tbl[i]["window"])

    wb.active.title = "raw data"
    sheet_raw = wb.get_sheet_by_name("raw data")

    titles = []
    if condition.find(CON_SORTING) != -1:
        titles = sorting_titles + titles
    for i in range(1, sheet_raw.max_column + 1):
        value = sheet_raw.cell(row=1, column=i).value
        if condition.find(CON_SORTING) != -1:
            if value not in sorting_titles:
                titles.append(value)
        else:
            titles.append(value)

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=20)
    bd = Side(style='thick', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    wb.add_named_style(highlight)

    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    sheet_cr = wb.create_sheet(index=0, title="CR list")
    sheet_cr.column_dimensions['A'].width = 26.57
    sheet_cr.cell(row=1, column=1).value = "Progress" + '\n' + "- Please update your current progress for this CR" + "\n" + "請說明目前的現況和進度"
    sheet_cr.cell(row=1, column=1).fill = yellow_fill
    sheet_cr.cell(row=1, column=1).border = border
    sheet_cr.column_dimensions['B'].width = 45.57
    sheet_cr.cell(row=1, column=2).value = "Actions" + '\n' + "- Please provide your actions for debugging this CR, and the expecting due date of each action" + '\n' + "請不要寫\"分析中\". 請列出接下來會作哪些事, 各自預計在什麼時間作完" + '\n' + "- Never just say you will transfer the CR to another colleague. Please sync with the next PIC to provide actions" + '\n' + "請不要寫\"我把CR轉給誰誰誰了\". 請跟下一手先串好, 列出接下來會作哪些事, 各自預計在什麼時間作完"
    sheet_cr.cell(row=1, column=2).fill = yellow_fill
    sheet_cr.cell(row=1, column=2).border = border
    for i in range(0, len(titles)):
        sheet_cr.cell(row=1, column=3+i).value = titles[i]
        sheet_cr.cell(row=1, column=3+i).fill = yellow_fill
        sheet_cr.cell(row=1, column=3+i).border = border
        if titles[i] == "id":
            sheet_cr.column_dimensions[sheet_cr.cell(row=1, column=3+i).column].width = 15
    for i in range(0, len(CRs)):
        sheet_cr.cell(row=2+i, column=1).border = border
        sheet_cr.cell(row=2+i, column=2).border = border
        for j in range(1, sheet_raw.max_column + 1):
            sheet_cr.cell(row=2+i, column=2+j).value = CRs[i][sheet_cr.cell(row=1, column=2+j).value]
            sheet_cr.cell(row=2+i, column=2+j).border = border
    filter = "A1:%s%d" % (get_column_letter(sheet_cr.max_column), sheet_cr.max_row)
    sheet_cr.auto_filter.ref = filter

    sheet_review = wb.create_sheet(index=1, title="review table")
    sheet_review.cell(row=2, column=2).value = "Session"
    sheet_review.cell(row=2, column=2).fill = yellow_fill
    sheet_review.cell(row=2, column=2).border = border
    sheet_review.column_dimensions['B'].width = 9.71
    sheet_review.cell(row=2, column=3).value = "Category"
    sheet_review.cell(row=2, column=3).fill = yellow_fill
    sheet_review.cell(row=2, column=3).border = border
    sheet_review.column_dimensions['C'].width = 9.71
    sheet_review.cell(row=2, column=4).value = "Dept"
    sheet_review.cell(row=2, column=4).fill = yellow_fill
    sheet_review.cell(row=2, column=4).border = border
    sheet_review.column_dimensions['D'].width = 20.43
    sheet_review.cell(row=2, column=5).value = "CR count"
    sheet_review.cell(row=2, column=5).fill = yellow_fill
    sheet_review.cell(row=2, column=5).border = border
    sheet_review.column_dimensions['E'].width = 8.14
    sheet_review.cell(row=2, column=6).value = "Attendee"
    sheet_review.cell(row=2, column=6).fill = yellow_fill
    sheet_review.cell(row=2, column=6).border = border
    sheet_review.column_dimensions['F'].width = 21.57
    sheet_review.cell(row=2, column=7).value = "Reviewed?"
    sheet_review.cell(row=2, column=7).fill = yellow_fill
    sheet_review.cell(row=2, column=7).border = border
    sheet_review.column_dimensions['G'].width = 10
    sheet_review.cell(row=2, column=8).value = "Manager"
    sheet_review.cell(row=2, column=8).fill = yellow_fill
    sheet_review.cell(row=2, column=8).border = border
    sheet_review.column_dimensions['H'].width = 23.71
    for i in range(0, len(review_tbl)):
        sheet_review.cell(row=3+i, column=3).value = review_tbl[i]["category"]
        sheet_review.cell(row=3+i, column=4).value = review_tbl[i]["team"]
        sheet_review.cell(row=3+i, column=5).value = review_tbl[i]["count"]
        sheet_review.cell(row=3+i, column=6).value = review_tbl[i]["window"]
        sheet_review.cell(row=3+i, column=8).value = review_tbl[i]["team"] + "_manager"
    filter = "A2:%s%d" % (get_column_letter(sheet_review.max_column), sheet_review.max_row + 1)
    sheet_review.auto_filter.ref = filter

    wb.active = 0
    wb.save(output_file)

    print("")
    attendees = ""
    for i in range(0, len(review_tbl)):
        if condition.find(CON_BYPASS_MODEM) != -1 and team_category(review_tbl[i]["team"]) == "Modem":
            continue
        if condition.find(CON_BYPASS_CONN) != -1 and team_category(review_tbl[i]["team"]) == "Conn":
            continue
        if condition.find(CON_BYPASS_QA) != -1 and team_category(review_tbl[i]["team"]) == "QA":
            continue
        if condition.find(CON_BYPASS_MM) != -1 and team_category(review_tbl[i]["team"]) == "MM":
            continue
        attendees += review_tbl[i]["window"] + ";"
    managers = ""
    for i in range(0, len(review_tbl)):
        team = review_tbl[i]["team"].upper()
        if condition.find(CON_BYPASS_MODEM) != -1 and team_category(team) == "Modem":
            continue
        if condition.find(CON_BYPASS_CONN) != -1 and team_category(team) == "Conn":
            continue
        if condition.find(CON_BYPASS_QA) != -1 and team_category(review_tbl[i]["team"]) == "QA":
            continue
        if condition.find(CON_BYPASS_MM) != -1 and team_category(review_tbl[i]["team"]) == "MM":
            continue
        if team.find("MBJ_") == -1 and team.find("MCD_") == -1 and team.find("MTI_") == -1 and team.find("MTB_") == -1  and team.find("MSZ_") == -1:
            team = "MTK_" + team
        managers += team + "_manager" + ";"
    print("Send the review meeting invitation to these windows:")
    print(attendees)
    print("")
    print("And CC to these managers:")
    print(managers)

if __name__ == "__main__":
    argv = sys.argv[1:]
    if len(argv) == 0:
        usage()
        sys.exit(0)

    try:
        opts, args = getopt.getopt(argv, "hi:o:m:c:", ["input=", "output=", "mapping=", "condition="])
    except getopt.GetoptError:
        usage()
        sys.exit(0)

    input_file = ""
    output_file = "CR_review_tbl.xlsx"
    mapping_file = ""
    condition = ""

    for opt, arg in opts:
        if opt == "-h":
            usage()
            sys.exit(0)
        elif opt in ("--input", "-i"):
            input_file = arg
        elif opt in ("--output", "-o"):
            output_file = arg
        elif opt in ("--mapping", "-m"):
            mapping_file = arg
        elif opt in ("--condition", "-c"):
            condition = arg

    if input_file == "":
        print("Error: input file is not given")
        usage()
        sys.exit(0)

    if mapping_file == "":
        print("Error: team/window mapping file is not given")
        usage()
        sys.exit(0)

    gen_CR_review_tbl(input_file, output_file, mapping_file, condition)
