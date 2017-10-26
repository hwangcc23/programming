#!/usr/bin/python
# This Python file uses the following encoding: utf-8

"""
review_rel_note_CR.py helps me automate a fucking boring stuff: find need-attention CRs
having some keywords in an excel file in which a CR list is provided for the release note
of S/W package.

Copyright (C) 2017 Chih-Chyuan Hwang (hwangcc@csie.nctu.edu.tw)

This program is free software; you can redistribute it and/or modify
it under the terms of the GNU General Public License version 2 as
published by the Free Software Foundation.
"""

import sys
import getopt
import openpyxl
import copy
import pyexcel # pip install pyexcel pyexcel-xls pyexcel-xlsx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def usage():
    print("review_rel_note_CR: process need-attention CRs in an excel file of a CR list for release note")
    print("Usage: rel_note_CR [options]")
    print("       options and arguments:")
    print("       -h: show help")
    print("       -i|--input FILENAME: give the input excel file")
    print("       -o|--output FILENAME: specify the output file name (default: \"__\"##input)")
    print("       -k|--keyword FILENAME: give the keyword file")
    print("       -a|--action ACTIONS: specify extra actions(\"keep_keyword_cr\")")

def get_keywords(keyword_file):
    keywords = []
    try:
        f = open(keyword_file, "r")
        try:
            while 1:
                line = f.readline()
                if line == "":
                    break
                if line[0] == "#":
                    continue
                keyword = line.strip('\n')
                if keyword == "":
                    continue
                keywords.append(copy.copy(keyword))
        except IOError:
            print("Fail to read " + keyword_file)
        finally:
            f.close()
    except IOError:
        print("Fail to open " + keyword_file)
    #print(keywords)
    return keywords

def mark_keywords(keywords, CR):
    need_attention = ""
    for key in CR:
        if CR[key] == None:
            if key != "eService ID" and need_attention.find("Empty") == -1:
                if need_attention != "":
                    need_attention = need_attention + ","
                need_attention = need_attention + "Empty"
            continue

        for i in range(0, len(keywords)):
            if CR[key].find(keywords[i]) != -1:
                if need_attention != "":
                    need_attention = need_attention + ","
                need_attention = need_attention + keywords[i].strip("\n")

    return need_attention

def review_rel_note_CR(input_file, output_file, keyword_file, actions):
    print("input_file = " + input_file, ", output_file = " + output_file, ", keyword_file = " + keyword_file, ", actions = " + actions)

    __, ext = input_file.rsplit(".", 1)
    if ext == "xls":
        orig_input_file = input_file
        input_file = input_file + "x"
        print("Convert " + orig_input_file + " to " + input_file)
        try:
            pyexcel.save_book_as(file_name=orig_input_file, dest_file_name=input_file)
        except IOError:
            print("Fail to convert " + orig_input_file)
            print("Abort")
            return

    print("Generating...")

    keywords = get_keywords(keyword_file)
    if len(keywords) == 0:
        print("No keyword is found")
        print("Abort")
        return

    try:
        wb = openpyxl.load_workbook(input_file)
    except IOError:
        print("Fail to load workbook from " + input_file)
        print("Abort")
        return

    sheet = wb.active

    Titles = []
    title_row = 1
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "CR ID":
            title_row = row
            break
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=title_row, column=col)
        Titles.append(copy.copy(cell.value))

    CRs = []
    for row in range(title_row + 1, sheet.max_row + 1):
        CR = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            CR[sheet.cell(row=title_row, column=col).value] = cell.value
        CRs.append(copy.copy(CR))

    # TODO: Add an option to remove CRs which info is empty

    reviewed_CRs = []
    for i in range(0, len(CRs)):
        need_attention = mark_keywords(keywords, CRs[i])
        if actions.find("keep_keyword_cr") != -1:
            CRs[i]["Need attention"] = need_attention
        elif need_attention != "":
            continue
        reviewed_CRs.append(copy.copy(CRs[i]))

    wrap_alignment = Alignment(wrap_text=True)
    font = Font(name='Arial', size=10)
    title_font = Font(name='Arial Black', size=10, bold=True)

    wb = Workbook()
    reviewed_sheet = wb.active

    for i in range(1, title_row):
        for j in range(1, sheet.max_column+1):
            c = reviewed_sheet.cell(row=i, column=j)
            c.value = sheet.cell(row=i, column=j).value
            #c.style = copy.copy(sheet.cell(row=i, column=j).style)
            c.alignment = copy.copy(sheet.cell(row=i, column=j).alignment)
            c.fill = copy.copy(sheet.cell(row=i, column=j).fill)
            c.font = copy.copy(sheet.cell(row=i, column=j).font)

    for i in range(0, len(Titles)):
        c = reviewed_sheet.cell(row=title_row, column=1+i)
        c.value = Titles[i]
        c.font = title_font
    if actions.find("keep_keyword_cr") != -1:
        print("keep_keyword_cr")
        reviewed_sheet.cell(row=title_row, column=1+len(Titles)).value = "Need attention"
        reviewed_sheet.cell(row=title_row, column=1+len(Titles)).font = title_font
    for i in range(0, len(reviewed_CRs)):
        for j in range(1, reviewed_sheet.max_column+1):
            c = reviewed_sheet.cell(row=title_row+1+i, column=j)
            c.value = reviewed_CRs[i][reviewed_sheet.cell(row=title_row, column=j).value]
            c.alignment = wrap_alignment
            c.font = font

    for i in range(title_row+1, sheet.max_row+1):
        reviewed_sheet.row_dimensions[i].height = 49.5
    reviewed_sheet.column_dimensions['A'].width = 15
    reviewed_sheet.column_dimensions['B'].width = 21
    reviewed_sheet.column_dimensions['C'].width = 60
    reviewed_sheet.column_dimensions['D'].width = 15
    reviewed_sheet.column_dimensions['E'].width = 15
    reviewed_sheet.column_dimensions['F'].width = 40
    reviewed_sheet.column_dimensions['G'].width = 15
    reviewed_sheet.column_dimensions['H'].width = 40
    reviewed_sheet.column_dimensions['I'].width = 40
    reviewed_sheet.column_dimensions['J'].width = 40
    reviewed_sheet.column_dimensions['K'].width = 40
    reviewed_sheet.column_dimensions['L'].width = 40
    reviewed_sheet.column_dimensions['M'].width = 15

    try:
        wb.save(output_file)
    except IOError:
        print("Fail to save " + output_file)
        print("Abort")
        return

    print("done")
    return

if __name__ == "__main__":
    argv = sys.argv[1:]
    if len(argv) == 0:
        usage()
        sys.exit(0)

    try:
        opts, args = getopt.getopt(argv, "hi:o:k:a:", ["input=", "output=", "keyword=", "action="])
    except getopt.GetoptError:
        usage()
        sys.exit(0)

    input_file = ""
    output_file = ""
    keyword_file = ""
    actions = ""

    for opt, arg in opts:
        if opt == "-h":
            usage()
            sys.exit(0)
        elif opt in ("--input", "-i"):
            input_file = arg
        elif opt in ("--output", "-o"):
            output_file = arg
        elif opt in ("--keyword", "-k"):
            keyword_file = arg
        elif opt in ("--action", "-a"):
            actions = arg

    if input_file == "":
        print("Error: input file is not given")
        usage()
        sys.exit(0)

    if output_file == "":
        output_file = "__" + input_file
        __, ext = output_file.rsplit(".", 1)
        if ext == "xls":
            output_file = output_file + "x"

    if keyword_file == "":
        print("Error: keyword file is not given")
        usage()
        sys.exit(0)

    review_rel_note_CR(input_file, output_file, keyword_file, actions)
